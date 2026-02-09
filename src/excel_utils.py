# src/excel_utils.py
from __future__ import annotations
import pandas as pd
from openpyxl import load_workbook
from typing import Dict, Optional

from .normalizer import normalize_text_spaces


def find_header_row_by_required_columns(df_preview: pd.DataFrame, required: list[str]) -> Optional[int]:
    """
    Busca la fila que contiene las columnas requeridas (como texto).
    Retorna índice de fila (0-based) para usar como header.
    """
    req_norm = [normalize_text_spaces(x) for x in required]
    for i in range(min(50, len(df_preview))):
        row = df_preview.iloc[i].astype(str).tolist()
        row_norm = [normalize_text_spaces(x) for x in row]
        if all(r in row_norm for r in req_norm):
            return i
    return None


def read_consolidado(path: str) -> pd.DataFrame:
    """
    Lee el consolidado aunque el header no esté en la primera fila.
    Requisito: columnas como CÓDIGO, AP. PATERNO, AP. MATERNO, NOMBRES.
    """
    preview = pd.read_excel(path, sheet_name=0, header=None, engine="openpyxl")
    header_row = find_header_row_by_required_columns(
        preview,
        required=["CÓDIGO", "AP. PATERNO", "AP. MATERNO", "NOMBRES"]
    )
    if header_row is None:
        raise ValueError("No se pudo detectar la fila de encabezados en el consolidado.")

    df = pd.read_excel(path, sheet_name=0, header=header_row, engine="openpyxl")
    df = df.loc[:, ~df.columns.astype(str).str.contains("^Unnamed")]
    return df


def load_template_catalogs(template_path: str) -> Dict[str, pd.DataFrame]:
    """
    Carga todas las hojas de la plantilla base (fija en resources/templates).
    """
    wb = load_workbook(template_path, data_only=True)
    catalogs: Dict[str, pd.DataFrame] = {}
    for name in wb.sheetnames:
        catalogs[name] = pd.read_excel(template_path, sheet_name=name, engine="openpyxl")
    return catalogs


def detect_columns(df: pd.DataFrame) -> Dict[str, Optional[str]]:
    """
    Mapea nombres "lógicos" a columnas reales del excel.
    ✅ Corrección clave: priorizar columna 'PAGADO' (amarilla BF) sobre 'PAGA'.
    """
    cols = {normalize_text_spaces(c): c for c in df.columns.astype(str)}

    def pick_exact(label: str) -> Optional[str]:
        key = normalize_text_spaces(label)
        return cols.get(key)

    def pick_contains(label: str) -> Optional[str]:
        key = normalize_text_spaces(label)
        for k, orig in cols.items():
            if key in k:
                return orig
        return None

    def pick(*labels: str) -> Optional[str]:
        # intenta exacto primero, luego contains
        for lb in labels:
            v = pick_exact(lb)
            if v:
                return v
        for lb in labels:
            v = pick_contains(lb)
            if v:
                return v
        return None

    # ✅ PAGO: prioridad estricta a PAGADO
    pagado_col = pick_exact("PAGADO") or pick_contains("PAGADO")
    # si no existe PAGADO, recién usar PAGA como último recurso
    if not pagado_col:
        pagado_col = pick_exact("PAGA") or pick_contains("PAGA") or pick("PAGO", "IMPORTE PAGADO")

    return {
        "codigo": pick("CÓDIGO", "CODIGO"),
        "ap_paterno": pick("AP. PATERNO", "AP PATERNO", "APELLIDO PATERNO"),
        "ap_materno": pick("AP. MATERNO", "AP MATERNO", "APELLIDO MATERNO"),
        "nombres": pick("NOMBRES", "NOMBRE"),
        "programa": pick("PROGRAMA ACADÉMICO", "PROGRAMA ACADEMICO"),
        "pension_escala": pick("PENSIÓN ESCALA", "PENSION ESCALA"),
        "ciclo": pick("CICLO"),
        "sede_filial": pick("SEDE/FILIAL", "SEDE FILIAL"),
        "tipo_doc": pick("TIPO DOCUMENTO", "TIPO DE DOCUMENTO"),
        "documento": pick("DOCUMENTO", "N° DOCUMENTO", "NRO DOCUMENTO"),
        "mail_personal": pick("MAIL PERSONAL", "CORREO PERSONAL", "EMAIL PERSONAL"),
        "telefonos": pick("TELEFONOS", "TELÉFONOS", "TELEFONO"),
        "direccion": pick("DIRECCION", "DIRECCIÓN"),
        "fecha_nac": pick("FECHA NAC.", "FECHA NAC", "FECHA NACIMIENTO"),
        "sexo": pick("SEXO"),
        "inicio": pick("INICIO"),
        "modalidad": pick("MODALIDAD MATRÍCULA", "MODALIDAD MATRICULA"),
        # ✅ aquí ya es BF (PAGADO) y no BE (PAGA)
        "pagado": pagado_col,
    }