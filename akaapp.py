# app.py
import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

from src.core import generate_outputs

TEMPLATE_PATH = "resources/templates/SubidaEstudiantesUAIFormato.xlsx"
SHEET_TARGET = "SubidaEstudiantes"

st.set_page_config(page_title="UAI - Generador de Usuarios", layout="wide")
st.title("🧩 Generador de Plantilla - Subida de Estudiantes (UAI)")

st.markdown("""
**Sube solo 2 archivos:**
1) **Consolidado Registro de Matrícula**  
2) **Export Outlook (usuarios existentes)**  

El sistema generará:
- ✅ Plantilla final lista para subir (solo columnas del template)
- 📋 Reporte de validación (incluye si ya tiene correo)
- ⚠️ Observados (con motivo)
""")

# =========================
# ✅ Estado (para que puedas re-cargar sin reiniciar el server)
# =========================
if "aprobados_df" not in st.session_state:
    st.session_state.aprobados_df = None
if "observados_df" not in st.session_state:
    st.session_state.observados_df = None

topbar1, topbar2 = st.columns([1, 1])
with topbar1:
    if st.button("🧹 Nueva carga / Limpiar resultados"):
        st.session_state.aprobados_df = None
        st.session_state.observados_df = None
        st.rerun()

col1, col2 = st.columns(2)
with col1:
    consolidado_file = st.file_uploader("📌 Consolidado Registro de Matrícula (.xlsx)", type=["xlsx"], key="consolidado")
with col2:
    outlook_file = st.file_uploader("📌 Export Outlook (xlsx o csv)", type=["xlsx", "csv"], key="outlook")

# =========================
# Helpers
# =========================
def _fix_arrow_types(df: pd.DataFrame) -> pd.DataFrame:
    """Evita crash de st.dataframe (pyarrow) cuando hay mezcla int/str como 'test'."""
    if df is None or df.empty:
        return df
    df2 = df.copy()

    if "TipoDeAdmission_RelationId" in df2.columns:
        df2["TipoDeAdmission_RelationId"] = df2["TipoDeAdmission_RelationId"].astype(str)

    # (opcional) si agregas más columnas mixtas en el futuro
    return df2

def build_template_output(template_path: str, aprobados: pd.DataFrame) -> bytes:
    """
    Exporta SOLO columnas del template (no agrega columnas nuevas),
    porque es el archivo que vas a subir al sistema UAI.
    """
    wb = load_workbook(template_path)
    if SHEET_TARGET not in wb.sheetnames:
        raise ValueError(f"No existe la hoja '{SHEET_TARGET}' en la plantilla.")

    ws = wb[SHEET_TARGET]

    headers = [c.value for c in ws[1]]
    headers = [h if h is not None else "" for h in headers]

    # limpiar filas existentes (desde fila 2)
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    # escribir datos respetando el orden de columnas de la plantilla
    records = aprobados.to_dict(orient="records")
    for i, row in enumerate(records, start=2):
        for j, h in enumerate(headers, start=1):
            if h in row:
                ws.cell(row=i, column=j, value=row[h])

    out = BytesIO()
    wb.save(out)
    return out.getvalue()

def build_excel_output(df: pd.DataFrame, sheet_name: str) -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return out.getvalue()

# =========================
# Acción principal
# =========================
if consolidado_file and outlook_file:
    if st.button("✅ Validar y Generar", type="primary"):
        with st.spinner("Procesando..."):
            import tempfile, os

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tf1:
                tf1.write(consolidado_file.getbuffer())
                consolidado_path = tf1.name

            suffix = ".csv" if outlook_file.name.lower().endswith(".csv") else ".xlsx"
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tf2:
                tf2.write(outlook_file.getbuffer())
                outlook_path = tf2.name

            try:
                aprobados_df, observados_df = generate_outputs(consolidado_path, outlook_path, TEMPLATE_PATH)
                # ✅ guardar en session_state
                st.session_state.aprobados_df = aprobados_df
                st.session_state.observados_df = observados_df
            finally:
                try:
                    os.unlink(consolidado_path)
                except:
                    pass
                try:
                    os.unlink(outlook_path)
                except:
                    pass

# =========================
# Mostrar resultados si existen
# =========================
aprobados_df = st.session_state.aprobados_df
observados_df = st.session_state.observados_df

if aprobados_df is not None and observados_df is not None:
    # ✅ Fix Arrow (pyarrow) antes de mostrar
    aprobados_view = _fix_arrow_types(aprobados_df)
    observados_view = _fix_arrow_types(observados_df)

    st.subheader("📊 Resultado")
    c1, c2, c3 = st.columns(3)
    c1.metric("Aprobados", int(len(aprobados_view)))
    c2.metric("Observados", int(len(observados_view)))

    # ✅ conteos de correo (si existen columnas)
    if "EstadoCorreoInstitucional" in aprobados_view.columns:
        ya_tiene = int((aprobados_view["EstadoCorreoInstitucional"] == "YA_TIENE").sum())
        generar = int((aprobados_view["EstadoCorreoInstitucional"] == "GENERAR").sum())
        c3.metric("Correo: YA_TIENE / GENERAR", f"{ya_tiene} / {generar}")
    else:
        c3.metric("Correo", "—")

    st.markdown("### ✅ Aprobados (preview)")
    st.dataframe(aprobados_view.head(50), width="stretch")

    st.markdown("### ⚠️ Observados (preview)")
    st.dataframe(observados_view.head(50), width="stretch")

    # =========================
    # Descargas
    # =========================
    try:
        # 1) Plantilla oficial para subir (solo columnas del template)
        plantilla_bytes = build_template_output(TEMPLATE_PATH, aprobados_view)

        # 2) Reporte de validación (incluye columnas nuevas)
        reporte_validacion_bytes = build_excel_output(aprobados_view, "Aprobados_Completo")

        # 3) Observados
        obs_bytes = build_excel_output(observados_view, "Observados")

        st.download_button(
            "⬇️ Descargar Plantilla Final (para subir)",
            data=plantilla_bytes,
            file_name="Plantilla_SubidaEstudiantes_GENERADA.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            "⬇️ Descargar Reporte de Validación (incluye correo/DNI)",
            data=reporte_validacion_bytes,
            file_name="Reporte_Validacion_Aprobados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        st.download_button(
            "⬇️ Descargar Observados (Motivos)",
            data=obs_bytes,
            file_name="Observados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"Error exportando: {e}")

else:
    st.info("Sube los 2 archivos para habilitar la generación.")